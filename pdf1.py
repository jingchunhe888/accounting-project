import pdfplumber
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from fuzzywuzzy import process, fuzz
from config import *
# Global variable to hold the count
_count = 0

def set_count(value):
    global _count
    _count = value

def get_count():
    global _count
    print(f'get was successfully set {type(_count)}')
    return _count

places = '''New York
Long Is City
Long Island C
Muellheim
Kifisia
San Francisco
Long Island C Astoria
LONG ISALND C ASTORIA
Long Island Long Island C
LONG ISLAND C LONG ISLAND C
Los Angeles
Long Island
Miami Beach
Koebenhavn
Las Vegas
Astoria
London
Seoul
Western Queens
San Luis Obis
South Beach
NEW YORK NEW YORK 
HENDERSON
WOODSIDE WOODSIDE
Toronto
Miami
CALIFON
NEWARK
LONG ISLAN LONG ISLAND C
Denville
ALBUQUERQUE
BROOKLYN
Columbus
Tucson 
Dallas
Washington
Washington DC
JAMAICA
'''

#this is for personal
def delete_places(description):

    global places
    place_list = []
    place_line = places.split('\n')
    for place in place_line:
        place_list.append(place)
    for place in place_list:
        pattern_place = re.compile(re.escape(place) + r' \b[A-Z]{2}\b',flags = re.IGNORECASE)
        match_place = re.search(pattern_place, description)
        if match_place:
            sub = match_place.group()
            description = re.sub(sub, '', description)
            return description


    pattern_end_two_upper = re.compile(r'\b[A-Z]{2}\b$')
    match_end_two_upper = re.search(pattern_end_two_upper,description)
    if match_end_two_upper:
        description = re.sub(match_end_two_upper,'')
        return description

    return description

# This is for credit card
# def delete_places(description):
#
#     global places
#     place_list = []
#     place_line = places.split('\n')
#     for place in place_line:
#         place_list.append(place)
#     for place in place_list:
#         pattern_place = re.compile(place + r' Card.*')
#         match_place = re.search(pattern_place, description)
#         if match_place:
#             description = re.sub(place + r' Card.*', '', description)
#             return description
#     return description




def clean_lines(input_line):
    date_or_check = ''
    pattern_date = re.compile(r'^\d{2}/\d{2}')
    pattern_check_number = re.compile(r'^\d{4}\b')  # Pattern to match four-digit number at the start of the line
    pattern_balance = re.compile(r'.*balance.*', flags=re.IGNORECASE)

    for index, char in enumerate(input_line):
        if pattern_balance.match(input_line):
            return None
        if char.isdigit() or char == '/':
            date_or_check += char
            match_date = pattern_date.match(date_or_check)
            match_check = pattern_check_number.match(date_or_check)
            if index == len(input_line)-5:
                #print('None')
                return None

            if match_date:
                date_or_check = match_date.group()
                end = index + 1
                remainingString = date_or_check + input_line[end:]
                #print('here')
                return remainingString
                    #print("remain" + remainingString)
                    #break
            if match_check:
                date_or_check = match_check.group()
                end = index + 1
                    # print(end)
                remainingString = date_or_check + input_line[end:]
                #print('here')
                return remainingString
                    # print("remain" + remainingString)
                    #break


    #print('nada')
    return None


def get_text(file_path):
    with pdfplumber.open(file_path) as pdf:
        lines = ''
        for page in pdf.pages:
            text = page.extract_text()
            lines = lines + text
            #print(text)
        #print("those are the lines" + lines)
        return lines.strip()

    #1






places = '''New York
Long Is City
Long Island C
Muellheim
Kifisia
San Francisco
Long Island C Astoria
LONG ISALND C ASTORIA
Long Island Long Island C
LONG ISLAND C LONG ISLAND C
Los Angeles
Long Island
Miami Beach
Koebenhavn
Las Vegas
Astoria
London
Seoul
Western Queens
San Luis Obis
South Beach
NEW YORK NEW YORK 
HENDERSON
WOODSIDE WOODSIDE
Toronto
Miami
CALIFON
NEWARK
LONG ISLAN LONG ISLAND C
Denville
ALBUQUERQUE
BROOKLYN
Columbus
Tucson 
Dallas
Washington
Washington DC
JAMAICA
'''

#this is for personal
def delete_places(description):

    global places
    place_list = []
    place_line = places.split('\n')
    for place in place_line:
        place_list.append(place)
    for place in place_list:
        pattern_place = re.compile(re.escape(place) + r' \b[A-Z]{2}\b') #took aware reignore case
        match_place = re.search(pattern_place, description)
        if match_place:
            print('match' + match_place.group())
            sub = match_place.group()
            description = re.sub(sub, '', description)
            return description
    return description

    pattern_end_two_upper = re.compile(r'\b[A-Z]{2}\b$')
    match_end_two_upper = re.search(pattern_end_two_upper,description)
    if match_end_two_upper:
        description = re.sub(match_end_two_upper,'')
        return description

# This is for credit card
# def delete_places(description):
#
#     global places
#     place_list = []
#     place_line = places.split('\n')
#     for place in place_line:
#         place_list.append(place)
#     for place in place_list:
#         pattern_place = re.compile(place + r' Card.*')
#         match_place = re.search(pattern_place, description)
#         if match_place:
#             description = re.sub(place + r' Card.*', '', description)
#             return description
#     return description




def clean_lines(input_line):
    date_or_check = ''
    pattern_date = re.compile(r'^\d{2}/\d{2}')
    pattern_check_number = re.compile(r'^\d{4}\b')  # Pattern to match four-digit number at the start of the line
    pattern_balance = re.compile(r'.*balance.*', flags=re.IGNORECASE)

    for index, char in enumerate(input_line):
        if pattern_balance.match(input_line):
            return None
        if char.isdigit() or char == '/':
            date_or_check += char
            match_date = pattern_date.match(date_or_check)
            match_check = pattern_check_number.match(date_or_check)
            if index == len(input_line)-5:
                #print('None')
                return None

            if match_date:
                date_or_check = match_date.group()
                end = index + 1
                remainingString = date_or_check + input_line[end:]
                #print('here')
                return remainingString
                    #print("remain" + remainingString)
                    #break
            if match_check:
                date_or_check = match_check.group()
                end = index + 1
                    # print(end)
                remainingString = date_or_check + input_line[end:]
                #print('here')
                return remainingString
                    # print("remain" + remainingString)
                    #break


    #print('nada')
    return None


def get_text(file_path):
    with pdfplumber.open(file_path) as pdf:
        lines = ''
        for page in pdf.pages:
            text = page.extract_text()
            lines = lines + text
            #print(text)
        #print("those are the lines" + lines)
        return lines.strip()

    #1


def filter_lines(input_text):
    # Define the patterns for MM/DD date and four-digit number
    pattern_date = re.compile(r'^\d{2}/\d{2}')  # Pattern to match MM/DD format at the start of the line
    pattern_date_middle = re.compile(r'.*\d{2}/\d{2}.*')  # Pattern to match MM/DD format at the start of the line
    pattern_check_number = re.compile(r'^\d{3,6}\b')  # Pattern to match four-digit number at the start of the line
    pattern_heading = re.compile(r'^\*start\*.*')
    pattern_end = re.compile(r'^\*end\*.*')
    pattern_end_deposits = re.compile(r'\*end\*deposits and additions', flags=re.IGNORECASE)
    pattern_start = re.compile(r'^\*start\*.*')
    pattern_positive = re.compile(r'DEPOSITS AND ADDITIONS', flags=re.IGNORECASE)
    pattern_negative = re.compile(r'Checks paid', flags=re.IGNORECASE)
    pattern_ending_balance = re.compile(r'.*start.*balance.*', flags=re.IGNORECASE)

    # Split the input text into lines
    lines = input_text.split('\n')

    # Filter lines that match either pattern at the start
    filtered_lines = []
    count = 0
    continue_count = True
    for line in lines:
        if pattern_end_deposits.match(line):
            continue_count = False
        # Check if the line matches the date pattern
        if pattern_date.match(line):
            filtered_lines.append(line)
            if continue_count:
                count += 1
            continue

        # Check if the line matches the check number pattern and has a date in the middle
        if pattern_check_number.match(line) and pattern_date_middle.search(line):
            filtered_lines.append(line)
            if continue_count:
                count += 1
            continue

        # Check if the line matches the heading pattern
        if pattern_heading.match(line):
            filtered_lines.append(line)
            continue

        # Check if the line matches the end pattern
        if pattern_end.match(line):
            filtered_lines.append(line)
            continue

    add_back = []
    for line in filtered_lines:
        if pattern_ending_balance.match(line):
            break
        if pattern_start.match(line):
            continue
        if pattern_end.match(line):
            keep = clean_lines(line)
            if keep:
                add_back.append(keep)
        else:
            add_back.append(line)

    # Join the filtered lines back into a single string
    filtered_text = '\n'.join(add_back)

    # Use the setter function to update the count
    set_count(count)
    print(f'this is the type of {type(count)}')

    return filtered_text


    # # Join the filtered lines back into a single string
    # filtered_text = '\n'.join(filtered_lines)
    # return filtered_text

# def main_pdf(file_path):
#     text = get_text(file_path)
#     text1 = filter_lines(text)
#     return text1

# def filter_people(text):
#     people = []
#     pattern_body =re.compile(r'Body Evolution')
#     pattern_william = re.compile(r'William Macagnone')
#     lines = text.split('\n')
#     for line in lines:
#         match_william = re.search(pattern_william, line)
#         match_body = re.search(pattern_body, line)
#
#         if match_william:
#             match_william = re.search(pattern_william, line).group()
#             people.append(match_william)
#         if match_body:
#             match_body = re.search(pattern_body, line).group()
#             people.append(match_body)
#     return people

def main_pdf(file_path):
    text = get_text(file_path)
    text = filter_lines(text)
    return text


    # Save the mismatches to a new Excel file (optional)
    #mismatches.to_excel('mismatches.xlsx', index=False)

# def get_people(file_path):
#     text = get_text(file_path)
#     text = filter_people(text)
#     text = pd.DataFrame(text, columns = ['People'])
#     return text

# text = get_text('/Users/jinhe/Downloads/2018/20181022-statements-7905-.pdf')
# print(text)

def get_missing_matches(file_path,sheet_name):
    #sheet_name = working_sheet_name
    working_df = pd.read_excel(file_path , sheet_name = sheet_name)

    # Filter rows where 'Account' is empty or NaN
    filtered_df = working_df[working_df['Account'].isna() | (working_df['Account'] == '')]

    # Remove duplicate descriptions
    filtered_df = filtered_df.drop_duplicates(subset=[subset_config])

    # Reset index for clean DataFrame
    filtered_df.reset_index(drop=True, inplace=True)
    filtered_df = filtered_df[headers_config]

    workbook = load_workbook(file_path)

    # Select the existing sheet
    worksheet = workbook[config_map]

    # Find the last row in the existing sheet
    last_row = worksheet.max_row

    # Append DataFrame to the existing sheet starting from the next row
    for row in dataframe_to_rows(filtered_df, index=False, header=False):
        worksheet.append(row)

    # Save the workbook
    workbook.save(file_path)
    # Load the updated 'Map' sheet into a DataFrame to access the newly added rows
    # updated_df = pd.read_excel(file_path, sheet_name='Map')
    # new_rows_df = updated_df.iloc[last_row:]

    print("Descriptions appended to the existing sheet.")

    # return new_rows_df
def fill_account(file_path,sheet_name):
    working_sheet_name = sheet_name  # Replace with the sheet name of your main Excel file
    map_sheet_name = config_map  # Replace with the sheet name of your secondary Excel file

    # Load the main Excel file with openpyxl to preserve formatting
    workbook_main = load_workbook(file_path)
    sheet_main = workbook_main[working_sheet_name]

    # Load the main DataFrame and secondary DataFrame with pandas
    df_main = pd.read_excel(file_path, sheet_name=working_sheet_name)
    df_secondary = pd.read_excel(file_path, sheet_name=map_sheet_name)

    # Create a dictionary from the secondary DataFrame with case-insensitive keys
    description_to_account = pd.Series(
        df_secondary.Account.values,
        index=df_secondary.Description.str.lower()
    ).to_dict()

    def fill(row):
        memo = row['Memo']
        if isinstance(memo, str) and memo.lower().startswith('recurring card purchase'):
            return 'Dues and Subscriptions'
        if pd.isna(row['Account']) or row['Account'] == '':
            try:
                return description_to_account.get(row['Description'].lower(), row['Account'])
            except Exception as e:
                print(description_to_account.get(row['Description']))

        return row['Account']

    # Apply the function to the main DataFrame using a lambda function
    df_main['Account'] = df_main.apply(lambda row: fill(row), axis=1)

    # Write back only the 'Account' column to the main Excel file using openpyxl
    for index, account in enumerate(df_main['Account'], start=2):  # Start from row 2, assuming headers are in row 1
        sheet_main.cell(row=index, column=4, value=account)  # Column 4 corresponds to 'Account'

    # Save the updated workbook
    workbook_main.save(file_path)

    print("Missing 'Account' values filled and updated in the main Excel file without altering other formatting.")

        # if pd.isna(row['Account']) or row['Account'] == '' or row['Account'].dtype == 'float' :
        #     return description_to_account.get(row['Description'].lower(), row['Account'])
        # if row['Memo'].lower().startswith('recurring card purchase'):
        #     return 'Dues and Subscriptions'
        # return row['Account']


def preprocess_text(text):
    """ Normalize text by converting to lowercase and removing special characters. """
    if isinstance(text, str):
        return re.sub(r'[^\w\s]', '', text.lower())
    else:
        return ''

def is_word_in_description(target, description):
    """ Check if every significant word in target is included in the description. """
    target_words = set(preprocess_text(target).split())
    description_words = set(preprocess_text(description).split())
    is_target = target_words.issubset(description_words)
    return target_words.issubset(description_words)

def all_words_in_description(messy_vendor, clean_vendors):
    """ Check if every word in clean_vendors is included in messy_vendor. """
    messy_words = set(re.findall(r'\w+', preprocess_text(messy_vendor)))
    best_match = None

    for clean_vendor in clean_vendors:
        clean_vendor_words = set(re.findall(r'\w+', preprocess_text(clean_vendor)))

        # Check if all words in clean vendor are in messy vendor
        if clean_vendor_words.issubset(messy_words):
            best_match = clean_vendor
            break  # Found a match, can exit the loop

    return best_match
def get_best_match(description, match_list):
    """ Get the best match from the match_list based on custom logic and fuzzy matching. """
    # Check for exact word inclusion first
    for match in match_list:
        if is_word_in_description(description, match):
            return match  # Return the match if it meets the word inclusion criteria
    best_match = all_words_in_description(description, match_list)
    if best_match:
        return best_match  # Return the match if it meets the word inclusion criteria

    # If no exact match is found, use fuzzy matching
    best_match, score = process.extractOne(description, match_list, scorer=fuzz.token_sort_ratio)
    return best_match if score >= 70 else None  # Adjust the threshold score as needed

def fill_closest_vendor(target_df, description_to_closest,sheet_name):
    # Create mask for rows where 'Account' is NaN or empty
    mask = target_df['Account'].isna() | (target_df['Account'] == '')
    # Update 'Closest Match' column in target_df only where the mask is True
    target_df.loc[mask, 'Closest Match'] = target_df.loc[mask, 'Description'].map(description_to_closest)
    # Save the updated DataFrame back to the Excel file
    with pd.ExcelWriter(file_path_config, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        target_df.to_excel(writer, sheet_name=sheet_name, index=False)  # Change 'TargetSheetName' to the actual sheet name
    print(f'Updated DataFrame:\n{target_df.head()}')

def get_closest_vendors(file_path,sheet_name):
    # Load the data from the Excel file
    df_map = pd.read_excel(file_path, sheet_name=sheet_name)

    # Filter DataFrame where 'Account' is NaN or empty
    filtered_df = df_map[df_map['Account'].isna() | (df_map['Account'] == '')].copy()

    # Create a list of descriptions from df_map where 'Account' is not NaN or empty
    closest_vendors_list = df_map.dropna(subset=['Account'])
    closest_vendors_list = closest_vendors_list[closest_vendors_list['Account'] != '']['Description'].tolist()

    # Apply custom matching for each row in the filtered DataFrame
    filtered_df['Closest Match'] = filtered_df['Description'].apply(lambda x: get_best_match(x, closest_vendors_list))
    filtered_df = filtered_df.dropna(subset=['Closest Match'])  # Drop rows where no match was found

    # Set 'Description' as the index for easier mapping
    filtered_df.set_index('Description', inplace=True)

    # Convert 'Closest Match' column to object type to avoid dtype issues
    filtered_df['Closest Match'] = filtered_df['Closest Match'].astype('object')

    # Load target DataFrame
    target_df = pd.read_excel(file_path, sheet_name=sheet_name)
    # Create a mapping from description to closest match
    description_to_closest = filtered_df['Closest Match'].to_dict()

    # Ensure 'Closest Match' column is of object type in target_df
    target_df['Closest Match'] = target_df['Closest Match'].astype('object')

    return target_df,description_to_closest
def fill_rules(file_path,rules_name,sheet_name):
    rulesdf = pd.read_excel(file_path,sheet_name=rules_name)
    rulesdf['Closest Match'] = rulesdf['Closest Match'].astype('object')
    rulesdf.drop_duplicates(subset=['Description'])
    rulesdf.set_index('Description',inplace=True)
    rules_dict = rulesdf['Closest Match'].to_dict()


    maindf = pd.read_excel(file_path, sheet_name=sheet_name)
    maindf['Description'] = maindf['Description'].replace(rules_dict)
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        maindf.to_excel(writer, sheet_name=sheet_name, index=False)
    print(f'Updated file saved as {file_path}')


def update_account(row):
    # Check if '!' is in the account column
    if pd.notna(row[account_config]) and '!' in row[account_config]:
        # Split the account value at '!'
        left_account, right_account = row[account_config].split('!')
        print(left_account)
        print(right_account)
        # Update account based on the presence of values in the credit and debit columns
        if pd.notna(row[positive_column_config]) and row[positive_column_config] != '':
            return right_account
        elif pd.notna(row[negative_column_config]) and row[negative_column_config] != '':
            return left_account
    return row['Account']


def split(file_path, sheet_name):
    # Load the data
    df = pd.read_excel(file_path, sheet_name)

    # Apply the update_account function
    df['Account'] = df.apply(update_account, axis=1)

    # Select the 'Account' column
    updated_df = df[['Account']]

    print(updated_df.to_string())
    print('columns')
    print(updated_df.columns)

    # Update the existing Excel sheet with the new 'Account' column
    update_account_column(file_path, sheet_name, updated_df)


def update_account_column(file_path, sheet_name, updated_df):
    # Load the existing workbook and sheet
    book = load_workbook(file_path)

    if sheet_name not in book.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in the file.")

    # Read the existing sheet into a DataFrame
    existing_df = pd.read_excel(file_path, sheet_name=sheet_name)

    # Make sure the 'Account' column is in the updated_df
    if 'Account' not in updated_df.columns:
        raise ValueError("The 'Account' column is missing from updated_df.")

    # Update the 'Account' column
    existing_df['Account'] = updated_df['Account']

    # Save the updated DataFrame back to the Excel sheet
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        existing_df.to_excel(writer, sheet_name=sheet_name, index=False)

    print(f'Updated file saved as {file_path}')

# Apply the function to the DataFrame


    # write code here
# debug text
# text = get_text('/Users/jinhe/Downloads/Different Bank Types/Chase Premier Plus Checking.pdf')
# text = filter_lines(text)
# print(text)
